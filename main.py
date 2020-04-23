print("\rLoading for Crowling ...",end='')
from bs4 import BeautifulSoup
import requests
import openpyxl
from time import time,sleep
from pprint import pprint
print("\rCrowling for withoutcert started !\n")
#num of lawyers 
header_name=['no','name in englis','name in chinese','link']
page_num=1
max_page_num=32
url='https://www.hklawsoc.org.hk/pub_e/memberlawlist/mem_withoutcert.asp?name=&pg={}&sj=0'
data=[]
frame=[]
while(page_num!=max_page_num):
    #   preparing for the current cycle
    sleep(3)
    print("\r[%0.5d] Get requist sent ..."%page_num,end='')
    response=requests.get(url.format(page_num),"")
    print("\r[%0.5d] Response received ..."%page_num,end='')
    rows=BeautifulSoup(response.text,"lxml")
    rows=rows.find_all("table")[10].findChildren("tr")[1:]
    for row in rows:
        col=row.findChildren("td")
        #   get lawyer no
        frame.append(col[0].getText().replace("\n",'').replace('\t','').replace('\r',''))
        #   get lawyer name in english
        frame.append(col[1].getText().replace("\n",'').replace('\t','').replace('\r',''))
        #   get lawyer name in chinese if found
        frame.append(col[2].getText().replace("\n",'').replace('\t','').replace('\r','').encode('utf8'))
        #   get lawyer link
        frame.append('https://www.hklawsoc.org.hk/pub_e/memberlawlist/'+col[1].a['href'])
        data.append(frame)
        frame=[]
    print("\r[%0.5d] Collecting this page finished ..."%page_num)
    page_num=page_num+1

print("Creating xl file ")
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
print("A %0.5d row to write to sheet"%(len(data)+2))
rownum=2
for lowyer in data:
    print("\r %0.5d"%(rownum),end='')
    #   no
    sheet.cell(row=rownum,column=1).value=lowyer[0]
    #   name english
    sheet.cell(row=rownum,column=2).value=lowyer[1]
    #   name chinese
    sheet.cell(row=rownum,column=3).value=lowyer[3]
    #   link
    #sheet.cell(row=rownum,column=4).value=lowyer[3]
    rownum=rownum+1
#   no
sheet.cell(row=1,column=1).value=header_name[0]
#   name english
sheet.cell(row=1,column=2).value=header_name[1]
#   name chinese
sheet.cell(row=1,column=3).value=header_name[3]#header_name[2]
#   link
#sheet.cell(row=1,column=4).value=






wb.save("withoutcert.xlsm")

print("\rwithoutcert xl is saved ...")


#######################################################################################################################




page_num=1
max_page_num=207
url='https://www.hklawsoc.org.hk/pub_e/memberlawlist/mem_withcert.asp?name=&pg={}&sj=0'
data=[]
frame=[]
print("\rCrowling for withoutcert started !\n")
while(page_num!=max_page_num):
    #   preparing for the current cycle
    sleep(3)
    print("\r[%0.5d] Get requist sent ..."%page_num,end='')
    response=requests.get(url.format(page_num),"")
    print("\r[%0.5d] Response received ..."%page_num,end='')
    rows=BeautifulSoup(response.text,"lxml")
    rows=rows.find_all("table")[10].findChildren("tr")[1:]
    for row in rows:
        col=row.findChildren("td")
        #   get lawyer no
        frame.append(col[0].getText().replace("\n",'').replace('\t','').replace('\r',''))
        #   get lawyer name in english
        frame.append(col[1].getText().replace("\n",'').replace('\t','').replace('\r',''))
        #   get lawyer name in chinese if found
        frame.append(col[2].getText().replace("\n",'').replace('\t','').replace('\r','').encode('utf8'))
        #   get lawyer link
        frame.append('https://www.hklawsoc.org.hk/pub_e/memberlawlist/'+col[1].a['href'])
        #print("\r[%0.5d]    %0.2d"%(page_num,int(frame[0])%50),end='')
        data.append(frame)
        frame=[]
    print("\r[%0.5d] Collecting this page finished ..."%page_num)
    page_num=page_num+1

print("Creating xl")
wb = openpyxl.Workbook()
sheet = wb.get_active_sheet()
print("A %0.5d row to write to sheet"%(len(data)+2))
rownum=2
#data[0][3]='Link'
for lowyer in data:
    print("\r %0.5d"%(rownum),end='')
    #   no
    sheet.cell(row=rownum,column=1).value=lowyer[0]
    #   name english
    sheet.cell(row=rownum,column=2).value=lowyer[1]
    #   name chinese
    sheet.cell(row=rownum,column=3).value=lowyer[3]
    #   link
    #sheet.cell(row=rownum,column=4).value=lowyer[3]
    rownum=rownum+1
#   no
sheet.cell(row=1,column=1).value=header_name[0]
#   name english
sheet.cell(row=1,column=2).value=header_name[1]
#   name chinese
sheet.cell(row=1,column=3).value=header_name[3]#header_name[2]
#   link
#sheet.cell(row=1,column=4).value=






wb.save("withcert.xlsm")

print("withcert is saved ...")